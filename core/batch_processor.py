import os
import glob
from lxml import etree
import openpyxl
from datetime import datetime

class BatchProcessor:
    # Diccionario inteligente para autodetectar columnas sin intervención humana.
    ALIAS_DICT = {
        "tercero": "CompanyID", # NIT
        "identificaci\u00f3n tercero": "CompanyID",
        "descripci": "ID",         # Número factura
        "fecha de elaboraci": "IssueDate", 
        "d\u00e9bito": "LineExtensionAmount",  # Base (Débito)
        "debito": "LineExtensionAmount", 
        "cr\u00e9dito": "PayableAmount",     # Total (Crédito)
        "credito": "PayableAmount",
        # Auto-relleno fijo según requerimiento del cliente:
        "tipo de comprobante": "_FIXED_CC",
        "sigla moneda": "_FIXED_COP",
        "tasa de cambio": "_FIXED_1",
        "sucursal": "_FIXED_0",
        "no. cuota": "_FIXED_1",
        "cuenta contable": "_ACCOUNT_DYNAMIC"
    }

    def __init__(self, template_path: str, debit_account: str = "", credit_account: str = ""):
        """
        Inicializa el procesador con la plantilla y las cuentas contables seleccionadas en la UI.
        """
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Plantilla no encontrada: {template_path}")
            
        self.template_path = template_path
        self.debit_account = debit_account
        self.credit_account = credit_account
        
        # Leemos los encabezados usando openpyxl para mapearlos a los tags
        wb = openpyxl.load_workbook(self.template_path, data_only=True)
        sheet = wb.active
        
        self.columns_order = []
        self.col_indices = {}
        
        # Asumimos que la fila 1 tiene los encabezados
        for cell in sheet[1]:
            if cell.value:
                col_name = str(cell.value).strip()
                self.columns_order.append(col_name)
                self.col_indices[col_name] = cell.column
                
        # Construimos el mapeo automático usando inteligencia de alias
        self.mapping = {}
        for col_name in self.columns_order:
            xml_tag_or_fixed = self._infer_tag(col_name)
            if xml_tag_or_fixed:
                self.mapping[col_name] = xml_tag_or_fixed
                
        # Guardamos referencias a las columnas Débito, Crédito y Cuenta para manipularlas
        self.debito_col = next((col for col in self.mapping if "debito" in col.lower() or "d\u00e9bito" in col.lower()), None)
        self.credito_col = next((col for col in self.mapping if "credito" in col.lower() or "cr\u00e9dito" in col.lower()), None)
        self.cuenta_col = next((col for col in self.mapping if self.mapping[col] == "_ACCOUNT_DYNAMIC"), None)
        
        # Namespaces comunes usados en la DIAN (UBL 2.1)
        self.namespaces = {
            'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
            'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
            'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2'
        }

    def _infer_tag(self, col_name: str) -> str:
        """Busca coincidencias lógicas entre el nombre de la columna y los campos XML o fijos"""
        col_lower = col_name.lower()
        for key, tag in self.ALIAS_DICT.items():
            if key in col_lower:
                return tag
        return None

    def _extract_numeric(self, str_value: str) -> float:
        """Helper para parsear texto a número de forma segura."""
        if not str_value:
            return 0.0
        try:
            if '.' in str_value and str_value.replace('.', '', 1).isdigit():
                return float(str_value)
            elif str_value.isdigit():
                return float(str_value)
            return 0.0
        except:
            return 0.0

    def _clean_nit(self, nit_str: str) -> str:
        """Limpia el NIT para que no tenga decimales ni letras raras (ej. 900.123.456.0 -> 900123456)"""
        if not nit_str: return ""
        # 1. Quitar todos los puntos, comas o guiones
        clean = nit_str.replace(".", "").replace(",", "").replace("-", "")
        # 2. Mantener solo los dígitos (limpia el DV si estaba separado, o lo unifica)
        clean = ''.join(c for c in clean if c.isdigit())
        return clean
        
    def _format_date_siigo(self, date_str: str) -> str:
        """Convierte fechas estándar (ej. YYYY-MM-DD) al formato DD/MM/AAAA para Siigo"""
        if not date_str: return ""
        try:
            # UBL suele traer YYYY-MM-DD
            d = datetime.strptime(date_str, "%Y-%m-%d")
            return d.strftime("%d/%m/%Y")
        except:
            return date_str

    def parse_single_xml_totals(self, xml_path: str) -> list:
        """
        Procesa el XML y extrae la información.
        Genera DOS filas para la contabilidad (Gasto y Pago).
        Aplica reglas de sanitización de NIT y Fechas.
        """
        try:
            tree = etree.parse(xml_path)
            root = tree.getroot()
            
            # --- 1. Extraer los datos compartidos (Cabecera) ---
            shared_data = {}
            for excel_col, xml_tag in self.mapping.items():
                
                # Manejar valores fijos primero
                if xml_tag == "_FIXED_CC":
                    shared_data[excel_col] = "CC"
                    continue
                elif xml_tag == "_FIXED_COP":
                    shared_data[excel_col] = "COP"
                    continue
                elif xml_tag == "_FIXED_1":
                    shared_data[excel_col] = 1
                    continue
                elif xml_tag == "_FIXED_0":
                    shared_data[excel_col] = 0
                    continue
                elif xml_tag == "_ACCOUNT_DYNAMIC":
                    # Lo llenaremos específicamente en cada fila
                    shared_data[excel_col] = "" 
                    continue
                
                # Búsqueda XPath para tags dinámicos
                xpath_expr = f"//*[local-name()='{xml_tag}']"
                elements = root.xpath(xpath_expr, namespaces=self.namespaces)
                
                value = None
                if elements:
                    for el in elements:
                        if el.text and el.text.strip():
                            value = el.text.strip()
                            break
                            
                # Sanitizaciones específicas
                if xml_tag == "CompanyID":
                    value = self._clean_nit(value)
                elif xml_tag == "IssueDate":
                    value = self._format_date_siigo(value)
                    
                shared_data[excel_col] = value

            # --- 2. Encontrar valores monetarios críticos ---
            base_val = 0.0
            if self.debito_col and self.debito_col in shared_data:
                base_val = self._extract_numeric(shared_data[self.debito_col])
                
            total_val = 0.0
            if self.credito_col and self.credito_col in shared_data:
                total_val = self._extract_numeric(shared_data[self.credito_col])

            # --- 3. Construir las 2 filas (2-row Double Entry para Siigo) ---
            rows_to_insert = []
            
            # Fila 1 (Gasto/Costo): Base al Débito, Crédito = 0, Cuenta Débito
            row_gasto = shared_data.copy()
            if self.debito_col: row_gasto[self.debito_col] = base_val
            if self.credito_col: row_gasto[self.credito_col] = 0.0
            if self.cuenta_col: row_gasto[self.cuenta_col] = self.debit_account
            rows_to_insert.append(row_gasto)
            
            # Fila 2 (Cuenta por Pagar): Crédito al Total, Débito = 0, Cuenta Crédito
            row_pago = shared_data.copy()
            if self.credito_col: row_pago[self.credito_col] = total_val
            if self.debito_col: row_pago[self.debito_col] = 0.0
            if self.cuenta_col: row_pago[self.cuenta_col] = self.credit_account
            rows_to_insert.append(row_pago)
                
            return rows_to_insert
            
        except Exception as e:
            print(f"\n[Error] falló el procesamiento de {xml_path}: {str(e)}")
            return []

    def process_folder(self, folder_path: str, output_excel_path: str):
        """
        Extrae data de los XMLs y los inserta en el Excel preservando diseño y formato.
        Añade 2 filas por comprobante asegurando la limpieza de Siigo.
        """
        search_pattern = os.path.join(folder_path, '*.xml')
        xml_files = glob.glob(search_pattern)
        
        total_files = len(xml_files)
        if total_files == 0:
            print(f"⚠️ No se encontraron archivos XML en: {folder_path}")
            return
            
        # 1. Cargamos el Excel original SIN romper formatos
        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active
        start_row = sheet.max_row + 1
        
        # 2. Procesamos y AÑADIMOS (append)
        for i, xml_file in enumerate(xml_files):
            data_rows = self.parse_single_xml_totals(xml_file)
            
            for data_row in data_rows:
                if data_row:
                    for col_name, value in data_row.items():
                        if col_name in self.col_indices:
                            col_idx = self.col_indices[col_name]
                            sheet.cell(row=start_row, column=col_idx, value=value)
                    
                    if "_Archivo_Origen" in self.col_indices:
                        col_idx = self.col_indices["_Archivo_Origen"]
                        sheet.cell(row=start_row, column=col_idx, value=os.path.basename(xml_file))
                        
                    start_row += 1
                
        # 3. Guardado final preservado
        wb.save(output_excel_path)
        print(f"\n✅ Archivo guardado correctamente en: {output_excel_path}")
